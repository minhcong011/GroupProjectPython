# Sửa câu hỏi
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .forms import LectureForm, BaiTapForm, CauHoiForm, CourseForm, CourseEditForm
from .models import BaiTap, CauHoi, Course, BaiLam, TestCase
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from core.models import Lecture, Account
from django.contrib import messages
from django.conf import settings

from django.utils import timezone
import json
import subprocess
import tempfile
import os

from django.db.models import Q


@login_required
def edit_question(request, question_id):
    question = get_object_or_404(CauHoi, id=question_id, bai_tap__nguoi_tao=request.user)
    if request.method == 'POST':
        form = CauHoiForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật câu hỏi thành công!')
            return redirect('edit_assignment', assignment_id=question.bai_tap.id)
    else:
        form = CauHoiForm(instance=question)
    return render(request, 'teacher_page/edit_question.html', {'form': form, 'question': question})

# Xóa câu hỏi
@login_required
def delete_question(request, question_id):
    question = get_object_or_404(CauHoi, id=question_id, bai_tap__nguoi_tao=request.user)
    assignment_id = question.bai_tap.id
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Đã xóa câu hỏi thành công!')
        return redirect('edit_assignment', assignment_id=assignment_id)
    return render(request, 'teacher_page/delete_question_confirm.html', {'question': question})



@login_required
def delete_assignment(request, assignment_id):
    try:
        baitap = BaiTap.objects.get(id=assignment_id, nguoi_tao=request.user)
    except BaiTap.DoesNotExist:
        return HttpResponseForbidden()
    if request.method == 'POST':
        baitap.delete()
        return redirect('assignment_list')
    return render(request, 'teacher_page/delete_assignment_confirm.html', {'assignment': baitap})

@login_required
def assignment_list(request):
    assignments = BaiTap.objects.filter(nguoi_tao=request.user)
    from core.models import Lecture
    lectures = Lecture.objects.filter(created_by=request.user)
    return render(request, 'teacher_page/assignment_list.html', {'assignments': assignments, 'lectures': lectures})


@login_required
def create_assignment(request):
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
        
    if request.method == 'POST':
        form = BaiTapForm(request.POST)
        if form.is_valid():
            baitap = form.save(commit=False)
            baitap.nguoi_tao = request.user
            baitap.save()
            if baitap.loai_baitap == 'quiz':
                return redirect('add_questions', assignment_id=baitap.id)
            return redirect('assignment_list')
        else:
            messages.error(request, 'Có lỗi trong form. Vui lòng kiểm tra lại.')
    else:
        form = BaiTapForm()
    return render(request, 'teacher_page/create_assignment.html', {'form': form})


@login_required
def add_questions(request, assignment_id):
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
        
    baitap = get_object_or_404(BaiTap, id=assignment_id, nguoi_tao=request.user)
    
    if request.method == 'POST':
        form = CauHoiForm(request.POST)
        if form.is_valid():
            cauhoi = form.save(commit=False)
            cauhoi.bai_tap = baitap
            cauhoi.save()
            messages.success(request, 'Thêm câu hỏi thành công!')
            if 'add_another' in request.POST:
                return redirect('add_questions', assignment_id=baitap.id)
            return redirect('assignment_list')
        else:
            messages.error(request, 'Có lỗi trong form câu hỏi. Vui lòng kiểm tra lại.')
    else:
        form = CauHoiForm()
    return render(request, 'teacher_page/add_questions.html', {'form': form, 'assignment': baitap})

# Create your views here.

# Sửa bài tập
from django.shortcuts import get_object_or_404
@login_required
def edit_assignment(request, assignment_id):
    assignment = get_object_or_404(BaiTap, id=assignment_id, nguoi_tao=request.user)
    if request.method == 'POST':
        form = BaiTapForm(request.POST, instance=assignment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật bài tập thành công!')
            return redirect('assignment_list')
    else:
        form = BaiTapForm(instance=assignment)
    return render(request, 'teacher_page/edit_assignment.html', {'form': form, 'assignment': assignment})
@login_required
def create_lecture(request):
    from core.models import Lecture
    courses = Course.objects.filter(created_by=request.user)
    if request.method == 'POST':
        form = LectureForm(request.POST, request.FILES)
        course_id = request.POST.get('course')
        selected_course = courses.filter(id=course_id).first() if course_id else None
        if form.is_valid() and selected_course:
            lecture = form.save(commit=False)
            lecture.created_by = request.user
            lecture.course = selected_course
            lecture.save()
            messages.success(request, 'Bài giảng đã được tạo thành công!')
            return redirect('lecture_list')
    else:
        form = LectureForm()
    return render(request, 'teacher_page/create_lecture.html', {'form': form, 'courses': courses})


# Xem chi tiết bài giảng (đề cương)
@login_required
def lecture_detail(request, pk):
    lecture = get_object_or_404(Lecture, pk=pk, created_by=request.user)
    return render(request, 'teacher_page/lecture_detail.html', {'lecture': lecture})

# Chỉnh sửa bài giảng (đề cương)
@login_required
def edit_lecture(request, pk):
    lecture = get_object_or_404(Lecture, pk=pk, created_by=request.user)
    if request.method == 'POST':
        form = LectureForm(request.POST, request.FILES, instance=lecture)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật đề cương thành công!')
            return redirect('lecture_list')
    else:
        form = LectureForm(instance=lecture)
    return render(request, 'teacher_page/edit_lecture.html', {'form': form, 'lecture': lecture})

# Xóa bài giảng (đề cương)
@login_required
def delete_lecture(request, pk):
    lecture = get_object_or_404(Lecture, pk=pk, created_by=request.user)
    if request.method == 'POST':
        lecture.delete()
        messages.success(request, 'Đã xóa đề cương thành công!')
        return redirect('lecture_list')
    return render(request, 'teacher_page/delete_lecture_confirm.html', {'lecture': lecture})
@login_required
def lecture_list(request):
    courses = Course.objects.filter(created_by=request.user)
    course_id = request.GET.get('course')
    selected_course = None
    from core.models import Lecture
    if course_id:
        selected_course = courses.filter(id=course_id).first()
        if selected_course:
            lectures = Lecture.objects.filter(course=selected_course)
        else:
            lectures = Lecture.objects.none()
    else:
        lectures = Lecture.objects.filter(course__in=courses)
    return render(request, 'teacher_page/lecture_list.html', {
        'lectures': lectures,
        'courses': courses,
        'selected_course': selected_course,
    })


@login_required
def course_management(request):
    query = request.GET.get('q', '')
    lang = request.GET.get('lang', '')

    # Lọc theo người dùng tạo
    courses = Course.objects.filter(created_by=request.user)

    # Tìm kiếm theo tên chứa từ khóa (không phân biệt hoa thường)
    if query:
        courses = courses.filter(name__icontains=query)

    # Lọc theo từ "Python" hoặc "Perl" xuất hiện trong tên
    if lang in ['Python', 'Perl']:
        courses = courses.filter(name__icontains=lang)

    form = CourseForm()
    context = {
        'courses': courses,
        'form': form,
        'query': query,
        'filter_lang': lang,
    }
    return render(request, 'teacher_page/course_management.html', context)


def create_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user  
            course.save()
            return redirect('course_management')
    return redirect('course_management')

def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if request.method == "POST":
        course.delete()
        return redirect('course_management')
    return redirect('course_management')

def edit_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        form = CourseEditForm(request.POST, instance=course)  
        if form.is_valid():
            updated = form.save(commit=False)
            updated.created_by = request.user
            updated.save()
            print(">>> Đã lưu thành công")
            return redirect('course_management')
        else:
            print(">>> Form không hợp lệ:", form.errors)
    else:
        form = CourseEditForm(instance=course)

    return render(request, 'teacher_page/edit_course.html', {'form': form, 'course': course})



@login_required
def cham_diem(request):
    """Trang chấm điểm"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    # Lấy danh sách bài tập
    bai_tap_list = BaiTap.objects.filter(nguoi_tao=request.user).order_by('-ngay_tao')
    
    # Lấy thống kê
    thong_ke = []
    for bai_tap in bai_tap_list:
        so_luong_nop = BaiLam.objects.filter(bai_tap=bai_tap).count()
        so_luong_cham = BaiLam.objects.filter(bai_tap=bai_tap, da_cham=True).count()
        
        thong_ke.append({
            'bai_tap': bai_tap,
            'so_luong_nop': so_luong_nop,
            'so_luong_cham': so_luong_cham,
            'chua_cham': so_luong_nop - so_luong_cham
        })
    
    context = {
        'thong_ke': thong_ke,
        'title': 'Chấm điểm bài tập'
    }
    return render(request, 'teacher_page/cham_diem.html', context)

@login_required
def chi_tiet_bai_tap(request, bai_tap_id):
    """Xem chi tiết bài làm của sinh viên"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    bai_tap = get_object_or_404(BaiTap, id=bai_tap_id, nguoi_tao=request.user)
    bai_lam_list = BaiLam.objects.filter(bai_tap=bai_tap).select_related('sinh_vien').order_by('-thoi_gian_nop')
    
    # Thêm tên file đã xử lý cho mỗi bài làm
    for bai_lam in bai_lam_list:
        if bai_lam.file_nop:
            import os
            bai_lam.file_name = os.path.basename(bai_lam.file_nop.name)
        else:
            bai_lam.file_name = None
    
    context = {
        'bai_tap': bai_tap,
        'bai_lam_list': bai_lam_list,
        'title': f'Chấm điểm: {bai_tap.tieu_de}'
    }
    return render(request, 'teacher_page/chi_tiet_bai_tap.html', context)

@login_required
def cham_bai_trac_nghiem(request, bai_lam_id):
    """Chấm bài trắc nghiệm tự động"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return JsonResponse({'success': False, 'message': 'Không có quyền truy cập'})
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không có quyền truy cập'})
    
    bai_lam = get_object_or_404(BaiLam, id=bai_lam_id)
    
    # Kiểm tra quyền
    if bai_lam.bai_tap.nguoi_tao != request.user:
        return JsonResponse({'success': False, 'message': 'Không có quyền chấm bài này'})
    
    # Chỉ chấm bài trắc nghiệm
    if bai_lam.bai_tap.loai_baitap != 'quiz':
        return JsonResponse({'success': False, 'message': 'Chỉ chấm được bài trắc nghiệm'})
    
    # Lấy đáp án đúng
    cau_hoi_list = CauHoi.objects.filter(bai_tap=bai_lam.bai_tap)
    dap_an_dung = {str(ch.id): ch.dap_an_dung for ch in cau_hoi_list}
    
    # Lấy đáp án sinh viên
    dap_an_sinh_vien = bai_lam.dap_an_json or {}
    
    # Tính điểm
    so_cau_dung = 0
    tong_so_cau = len(dap_an_dung)
    
    for cau_id, dap_an_dung_val in dap_an_dung.items():
        if dap_an_sinh_vien.get(cau_id) == dap_an_dung_val:
            so_cau_dung += 1
    
    diem_so = (so_cau_dung / tong_so_cau) * 10 if tong_so_cau > 0 else 0
    
    # Lưu điểm
    bai_lam.diem_so = round(diem_so, 2)
    bai_lam.da_cham = True
    bai_lam.so_cau_dung = so_cau_dung
    bai_lam.tong_so_cau = tong_so_cau
    bai_lam.save()
    
    return JsonResponse({
        'success': True,
        'diem_so': bai_lam.diem_so,
        'so_cau_dung': so_cau_dung,
        'tong_so_cau': tong_so_cau
    })

@login_required
def cham_bai_lap_trinh(request, bai_lam_id):
    """Chấm bài lập trình bằng test case"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return JsonResponse({'success': False, 'message': 'Không có quyền truy cập'})
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không có quyền truy cập'})
    
    bai_lam = get_object_or_404(BaiLam, id=bai_lam_id)
    
    # Kiểm tra quyền
    if bai_lam.bai_tap.nguoi_tao != request.user:
        return JsonResponse({'success': False, 'message': 'Không có quyền chấm bài này'})
    
    # Chỉ chấm bài lập trình
    if bai_lam.bai_tap.loai_baitap != 'code':
        return JsonResponse({'success': False, 'message': 'Chỉ chấm được bài lập trình'})
    
    # Lấy test cases
    test_cases = TestCase.objects.filter(bai_tap=bai_lam.bai_tap)
    
    if not test_cases.exists():
        return JsonResponse({'success': False, 'message': 'Chưa có test case cho bài này'})
    
    # Chạy test
    ket_qua_test = []
    tong_diem = 0
    so_test_pass = 0
    diem_toi_da = sum([tc.diem_so for tc in test_cases])
    
    for test_case in test_cases:
        try:
            # Tạo file tạm để chạy code với encoding UTF-8
            with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as f:
                # Thêm encoding declaration và sửa code để tương thích với tiếng Việt
                code_to_run = f"""# -*- coding: utf-8 -*-
import sys
import os
import locale

# Set UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')
os.environ['PYTHONIOENCODING'] = 'utf-8'

# Set locale to support Vietnamese
try:
    locale.setlocale(locale.LC_ALL, 'vi_VN.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Vietnamese_Vietnam.1258')
    except:
        pass

{bai_lam.code_nop}
"""
                f.write(code_to_run)
                temp_file = f.name
            
            # Chạy code với input và environment UTF-8 cho tiếng Việt
            env = os.environ.copy()
            env['PYTHONIOENCODING'] = 'utf-8'
            env['LANG'] = 'vi_VN.UTF-8'
            env['LC_ALL'] = 'vi_VN.UTF-8'
            
            process = subprocess.Popen(
                ['python', temp_file],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                env=env,
                encoding='utf-8'
            )
            
            try:
                stdout, stderr = process.communicate(input=test_case.input_data, timeout=5)
            except subprocess.TimeoutExpired:
                process.kill()
                raise subprocess.TimeoutExpired(process.args, 5)
            
            # So sánh output (Vietnamese-aware comparison)
            output_sinh_vien = stdout.strip()
            expected_output = test_case.expected_output.strip()
            
            # Import Vietnamese helper functions
            import unicodedata
            import re
            
            def normalize_vietnamese_text(text):
                if not text:
                    return ""
                normalized = unicodedata.normalize('NFC', text)
                return normalized.strip()
            
            def compare_vietnamese_output(actual, expected):
                actual_norm = normalize_vietnamese_text(actual)
                expected_norm = normalize_vietnamese_text(expected)
                
                # Direct comparison first
                if actual_norm == expected_norm:
                    return True, actual_norm, expected_norm
                
                # Try case-insensitive comparison
                if actual_norm.lower() == expected_norm.lower():
                    return True, actual_norm, expected_norm
                
                # Try extracting numbers for flexible comparison
                actual_numbers = re.findall(r'-?\d+', actual_norm)
                expected_numbers = re.findall(r'-?\d+', expected_norm)
                
                if actual_numbers and expected_numbers:
                    if actual_numbers[-1] == expected_numbers[-1]:
                        return True, actual_numbers[-1], expected_numbers[-1]
                
                return False, actual_norm, expected_norm
            
            # Use Vietnamese-aware comparison
            is_match, processed_actual, processed_expected = compare_vietnamese_output(
                output_sinh_vien, expected_output
            )
            
            # Update values for final comparison
            if is_match and processed_actual != output_sinh_vien:
                output_sinh_vien = processed_actual
                expected_output = processed_expected
            
            if output_sinh_vien == expected_output:
                ket_qua = 'PASS'
                diem_dat = test_case.diem_so
                tong_diem += diem_dat
                so_test_pass += 1
            else:
                ket_qua = 'FAIL'
                diem_dat = 0
            
            ket_qua_test.append({
                'ten_test': test_case.ten_test,
                'input': test_case.input_data,
                'expected': expected_output,
                'actual': output_sinh_vien,
                'ket_qua': ket_qua,
                'diem_dat': diem_dat,
                'diem_toi_da': test_case.diem_so,
                'error': stderr if stderr else None,
                'original_output': stdout.strip()  # Giữ lại output gốc để debug
            })
            
            # Xóa file tạm
            os.unlink(temp_file)
            
        except subprocess.TimeoutExpired:
            ket_qua_test.append({
                'ten_test': test_case.ten_test,
                'input': test_case.input_data,
                'expected': test_case.expected_output,
                'actual': '',
                'ket_qua': 'TIMEOUT',
                'diem_dat': 0,
                'diem_toi_da': test_case.diem_so,
                'error': 'Code chạy quá lâu (timeout)'
            })
            if 'temp_file' in locals():
                os.unlink(temp_file)
            
        except Exception as e:
            ket_qua_test.append({
                'ten_test': test_case.ten_test,
                'input': test_case.input_data,
                'expected': test_case.expected_output,
                'actual': '',
                'ket_qua': 'ERROR',
                'diem_dat': 0,
                'diem_toi_da': test_case.diem_so,
                'error': str(e)
            })
            if 'temp_file' in locals():
                os.unlink(temp_file)
    
    # Tính điểm cuối cùng (thang điểm 10)
    diem_cuoi = (tong_diem / diem_toi_da) * 10 if diem_toi_da > 0 else 0
    
    # Lưu kết quả
    bai_lam.diem_so = round(diem_cuoi, 2)
    bai_lam.da_cham = True
    bai_lam.ket_qua_test = ket_qua_test
    bai_lam.so_test_pass = so_test_pass
    bai_lam.tong_so_test = len(test_cases)
    bai_lam.save()
    
    return JsonResponse({
        'success': True,
        'diem_so': bai_lam.diem_so,
        'ket_qua_test': ket_qua_test,
        'tong_diem': tong_diem,
        'diem_toi_da': diem_toi_da
    })

@login_required
def quan_ly_test_case(request, bai_tap_id):
    """Quản lý test case cho bài lập trình"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    bai_tap = get_object_or_404(BaiTap, id=bai_tap_id, nguoi_tao=request.user)
    
    if bai_tap.loai_baitap != 'code':
        messages.error(request, 'Chỉ có thể tạo test case cho bài lập trình!')
        return redirect('cham_diem')
    
    if request.method == 'POST':
        ten_test = request.POST.get('ten_test')
        input_data = request.POST.get('input_data')
        expected_output = request.POST.get('expected_output')
        diem_so = float(request.POST.get('diem_so', 1.0))
        
        TestCase.objects.create(
            bai_tap=bai_tap,
            ten_test=ten_test,
            input_data=input_data,
            expected_output=expected_output,
            diem_so=diem_so
        )
        
        messages.success(request, 'Thêm test case thành công!')
        return redirect('quan_ly_test_case', bai_tap_id=bai_tap_id)
    
    test_cases = TestCase.objects.filter(bai_tap=bai_tap)
    
    context = {
        'bai_tap': bai_tap,
        'test_cases': test_cases,
        'title': f'Test Case: {bai_tap.tieu_de}'
    }
    return render(request, 'teacher_page/quan_ly_test_case.html', context)

@login_required
def debug_csrf(request):
    """Debug view để test CSRF token"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    if request.method == 'POST':
        messages.success(request, 'CSRF test thành công!')
    
    return render(request, 'teacher_page/debug_csrf.html')

@login_required
def test_create_assignment(request):
    """Test view để tạo bài tập đơn giản"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    if request.method == 'POST':
        try:
            tieu_de = request.POST.get('tieu_de')
            mo_ta = request.POST.get('mo_ta')
            loai_baitap = request.POST.get('loai_baitap')
            han_nop = request.POST.get('han_nop')
            
            # Validate data
            if not all([tieu_de, loai_baitap, han_nop]):
                messages.error(request, 'Vui lòng điền đầy đủ thông tin!')
                return render(request, 'teacher_page/test_create_assignment.html')
            
            # Parse datetime
            from datetime import datetime
            han_nop_datetime = datetime.fromisoformat(han_nop.replace('T', ' '))
            
            # Create assignment
            baitap = BaiTap.objects.create(
                tieu_de=tieu_de,
                mo_ta=mo_ta,
                loai_baitap=loai_baitap,
                han_nop=han_nop_datetime,
                nguoi_tao=request.user
            )
            
            messages.success(request, f'Tạo bài tập "{tieu_de}" thành công!')
            
            if loai_baitap == 'quiz':
                return redirect('add_questions', assignment_id=baitap.id)
            else:
                return redirect('assignment_list')
                
        except Exception as e:
            messages.error(request, f'Lỗi khi tạo bài tập: {str(e)}')
    
    return render(request, 'teacher_page/test_create_assignment.html')

@login_required
def tao_test_case_tu_dong(request, bai_tap_id):
    """Tạo test cases tự động cho bài lập trình"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    bai_tap = get_object_or_404(BaiTap, id=bai_tap_id, nguoi_tao=request.user)
    
    if bai_tap.loai_baitap != 'code':
        messages.error(request, 'Chỉ có thể tạo test case cho bài lập trình!')
        return redirect('quan_ly_test_case', bai_tap_id=bai_tap_id)
    
    # Xóa test cases cũ nếu có
    TestCase.objects.filter(bai_tap=bai_tap).delete()
    
    # Tạo test cases mẫu dựa trên tên bài tập (Vietnamese-friendly)
    if 'vị trí' in bai_tap.tieu_de.lower() or 'vi tri' in bai_tap.tieu_de.lower():
        # Test cases cho bài "Tìm vị trí phần tử âm đầu tiên" - Flexible với tiếng Việt
        test_cases_data = [
            {
                'ten_test': 'Test 1: Có phần tử âm ở vị trí 1',
                'input_data': '',
                'expected_output': '1',  # Chỉ số, sẽ match với "Vị trí phần tử âm đầu tiên là: 1"
                'diem_so': 2.5
            },
            {
                'ten_test': 'Test 2: Phần tử âm ở đầu',
                'input_data': '',
                'expected_output': '0',  # Sẽ match với "Vị trí phần tử âm đầu tiên là: 0"
                'diem_so': 2.5
            },
            {
                'ten_test': 'Test 3: Không có phần tử âm',
                'input_data': '',
                'expected_output': '-1',  # Sẽ cần xử lý đặc biệt cho câu "Không có..."
                'diem_so': 2.5
            },
            {
                'ten_test': 'Test 4: Nhiều phần tử âm',
                'input_data': '',
                'expected_output': '2',  # Lấy vị trí đầu tiên
                'diem_so': 2.5
            }
        ]
    elif 'số nguyên tố' in bai_tap.tieu_de.lower():
        # Test cases cho bài kiểm tra số nguyên tố
        test_cases_data = [
            {
                'ten_test': 'Test 1: Số nguyên tố',
                'input_data': '7',
                'expected_output': 'YES',
                'diem_so': 2.5
            },
            {
                'ten_test': 'Test 2: Không phải số nguyên tố',
                'input_data': '8',
                'expected_output': 'NO',
                'diem_so': 2.5
            },
            {
                'ten_test': 'Test 3: Số 1',
                'input_data': '1',
                'expected_output': 'NO',
                'diem_so': 2.5
            },
            {
                'ten_test': 'Test 4: Số 2',
                'input_data': '2',
                'expected_output': 'YES',
                'diem_so': 2.5
            }
        ]
    else:
        # Test cases mặc định
        test_cases_data = [
            {
                'ten_test': 'Test 1: Input cơ bản',
                'input_data': '5',
                'expected_output': '5',
                'diem_so': 5.0
            },
            {
                'ten_test': 'Test 2: Input đặc biệt',
                'input_data': '0',
                'expected_output': '0',
                'diem_so': 5.0
            }
        ]
    
    # Tạo test cases
    for data in test_cases_data:
        TestCase.objects.create(
            bai_tap=bai_tap,
            ten_test=data['ten_test'],
            input_data=data['input_data'],
            expected_output=data['expected_output'],
            diem_so=data['diem_so']
        )
    
    messages.success(request, f'Đã tạo {len(test_cases_data)} test cases tự động!')
    return redirect('quan_ly_test_case', bai_tap_id=bai_tap_id)

@login_required
def xoa_test_case(request, test_case_id):
    """Xóa một test case"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    test_case = get_object_or_404(TestCase, id=test_case_id, bai_tap__nguoi_tao=request.user)
    bai_tap_id = test_case.bai_tap.id
    
    if request.method == 'POST':
        test_case.delete()
        messages.success(request, 'Đã xóa test case!')
        return redirect('quan_ly_test_case', bai_tap_id=bai_tap_id)
    
    return render(request, 'teacher_page/confirm_delete_test_case.html', {
        'test_case': test_case,
        'bai_tap': test_case.bai_tap
    })

@login_required
def ket_qua_cham_diem(request, bai_lam_id):
    """Xem kết quả chấm điểm chi tiết"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return redirect('student_home')
    except Account.DoesNotExist:
        return redirect('signin')
    
    bai_lam = get_object_or_404(BaiLam, id=bai_lam_id)
    
    # Kiểm tra quyền
    if bai_lam.bai_tap.nguoi_tao != request.user:
        messages.error(request, 'Không có quyền xem bài này!')
        return redirect('cham_diem')
    
    # Thêm tên file đã xử lý
    if bai_lam.file_nop:
        import os
        bai_lam.file_name = os.path.basename(bai_lam.file_nop.name)
    else:
        bai_lam.file_name = None
    
    context = {
        'bai_lam': bai_lam,
        'title': f'Kết quả chấm điểm: {bai_lam.sinh_vien.username}'
    }
    return render(request, 'teacher_page/ket_qua_cham_diem.html', context)

@login_required
def sua_diem(request, bai_lam_id):
    """Sửa điểm thủ công"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return JsonResponse({'success': False, 'message': 'Không có quyền truy cập'})
    except Account.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Không có quyền truy cập'})
    
    bai_lam = get_object_or_404(BaiLam, id=bai_lam_id)
    
    # Kiểm tra quyền
    if bai_lam.bai_tap.nguoi_tao != request.user:
        return JsonResponse({'success': False, 'message': 'Không có quyền sửa bài này'})
    
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        diem_so = data.get('diem_so')
        nhan_xet = data.get('nhan_xet', '')
        
        if diem_so is None or not (0 <= diem_so <= 10):
            return JsonResponse({'success': False, 'message': 'Điểm phải từ 0 đến 10'})
        
        bai_lam.diem_so = diem_so
        bai_lam.nhan_xet = nhan_xet
        bai_lam.da_cham = True
        bai_lam.save()
        
        return JsonResponse({'success': True, 'diem_so': diem_so})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@login_required
def download_file(request, bai_lam_id):
    """Download file nộp của sinh viên"""
    try:
        account = Account.objects.get(username=request.user.username)
        if not account.is_teacher:
            return HttpResponseForbidden('Không có quyền truy cập')
    except Account.DoesNotExist:
        return HttpResponseForbidden('Không có quyền truy cập')
    
    bai_lam = get_object_or_404(BaiLam, id=bai_lam_id)
    
    # Kiểm tra quyền - chỉ giáo viên tạo bài tập mới được download
    if bai_lam.bai_tap.nguoi_tao != request.user:
        return HttpResponseForbidden('Không có quyền download file này')
    
    # Kiểm tra xem có file không
    if not bai_lam.file_nop:
        return HttpResponse('Sinh viên chưa nộp file', status=404)
    
    # Tạo response với file
    try:
        import os
        import mimetypes
        
        file_path = bai_lam.file_nop.path
        file_name = os.path.basename(bai_lam.file_nop.name)
        
        # Xác định content type
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    except FileNotFoundError:
        return HttpResponse('File không tồn tại trên server', status=404)
    except Exception as e:
        return HttpResponse(f'Lỗi khi tải file: {str(e)}', status=500)


from django.db.models import Count, Avg, Q
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
import csv
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO

@login_required
def thong_ke_bao_cao(request):
    """Trang thống kê và báo cáo cho giảng viên"""
    user = request.user
    
    # Lấy tất cả bài tập của giảng viên
    bai_taps = BaiTap.objects.filter(nguoi_tao=user)
    
    # Thống kê tổng quan
    tong_bai_tap = bai_taps.count()
    tong_bai_nop = BaiLam.objects.filter(bai_tap__nguoi_tao=user).count()
    bai_da_cham = BaiLam.objects.filter(bai_tap__nguoi_tao=user, da_cham=True).count()
    
    # Tính điểm trung bình
    diem_trung_binh = BaiLam.objects.filter(
        bai_tap__nguoi_tao=user, 
        da_cham=True,
        diem_so__isnull=False
    ).aggregate(avg_score=Avg('diem_so'))['avg_score'] or 0
    
    # Thống kê theo tháng (6 tháng gần nhất)
    now = timezone.now()
    six_months_ago = now - timedelta(days=180)
    
    monthly_stats = BaiLam.objects.filter(
        bai_tap__nguoi_tao=user,
        thoi_gian_nop__gte=six_months_ago
    ).annotate(
        month=TruncMonth('thoi_gian_nop')
    ).values('month').annotate(
        count=Count('id'),
        avg_score=Avg('diem_so', filter=Q(da_cham=True))
    ).order_by('month')
    
    # Thống kê theo bài tập
    bai_tap_stats = []
    for bai_tap in bai_taps:
        bai_lams = BaiLam.objects.filter(bai_tap=bai_tap)
        da_cham = bai_lams.filter(da_cham=True)
        
        avg_score = da_cham.aggregate(avg=Avg('diem_so'))['avg'] or 0
        
        bai_tap_stats.append({
            'bai_tap': bai_tap,
            'tong_nop': bai_lams.count(),
            'da_cham': da_cham.count(),
            'chua_cham': bai_lams.filter(da_cham=False).count(),
            'diem_tb': round(avg_score, 1) if avg_score else 0,
            'ti_le_hoan_thanh': round((da_cham.count() / bai_lams.count() * 100), 1) if bai_lams.count() > 0 else 0
        })
    
    # Thống kê sinh viên
    sinh_vien_stats = BaiLam.objects.filter(
        bai_tap__nguoi_tao=user
    ).values(
        'sinh_vien__username',
        'sinh_vien__first_name', 
        'sinh_vien__last_name'
    ).annotate(
        tong_bai_nop=Count('id'),
        bai_da_cham=Count('id', filter=Q(da_cham=True)),
        diem_tb=Avg('diem_so', filter=Q(da_cham=True))
    ).order_by('-diem_tb')
    
    context = {
        'title': 'Thống kê & Báo cáo',
        'tong_bai_tap': tong_bai_tap,
        'tong_bai_nop': tong_bai_nop,
        'bai_da_cham': bai_da_cham,
        'bai_chua_cham': tong_bai_nop - bai_da_cham,
        'diem_trung_binh': round(diem_trung_binh, 1),
        'monthly_stats': list(monthly_stats),
        'bai_tap_stats': bai_tap_stats,
        'sinh_vien_stats': sinh_vien_stats,
        'now': now,
    }
    
    return render(request, 'teacher_page/thong_ke_bao_cao.html', context)

@login_required
def export_excel(request):
    """Xuất báo cáo Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        # Tạo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Thống kê bài tập"
        
        # Header
        ws['A1'] = 'BÁO CÁO THỐNG KÊ BÀI TẬP'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:G1')
        
        ws['A3'] = f'Giảng viên: {request.user.get_full_name() or request.user.username}'
        ws['A4'] = f'Ngày xuất: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Headers cho bảng
        headers = ['STT', 'Tên bài tập', 'Loại', 'Tổng nộp', 'Đã chấm', 'Chưa chấm', 'Điểm TB']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dữ liệu
        bai_taps = BaiTap.objects.filter(nguoi_tao=request.user)
        for row, bai_tap in enumerate(bai_taps, 7):
            bai_lams = BaiLam.objects.filter(bai_tap=bai_tap)
            da_cham = bai_lams.filter(da_cham=True)
            avg_score = da_cham.aggregate(avg=Avg('diem_so'))['avg'] or 0
            
            ws.cell(row=row, column=1, value=row-6)
            ws.cell(row=row, column=2, value=bai_tap.tieu_de)
            ws.cell(row=row, column=3, value='Lập trình' if bai_tap.loai_baitap == 'code' else 'Trắc nghiệm')
            ws.cell(row=row, column=4, value=bai_lams.count())
            ws.cell(row=row, column=5, value=da_cham.count())
            ws.cell(row=row, column=6, value=bai_lams.filter(da_cham=False).count())
            ws.cell(row=row, column=7, value=round(avg_score, 1) if avg_score else 0)
        
        # Tạo response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="thong_ke_bai_tap_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response
        
    except ImportError:
        messages.error(request, 'Chưa cài đặt openpyxl. Vui lòng cài đặt: pip install openpyxl')
        return redirect('thong_ke_bao_cao')
    except Exception as e:
        messages.error(request, f'Lỗi xuất Excel: {str(e)}')
        return redirect('thong_ke_bao_cao')

@login_required  
def export_pdf(request):
    """Xuất báo cáo PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.units import inch
        
        # Tạo response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="bao_cao_thong_ke_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        # Tạo PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Tiêu đề
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredText(width/2, height-50, "BÁO CÁO THỐNG KÊ BÀI TẬP")
        
        # Thông tin
        p.setFont("Helvetica", 12)
        p.drawString(50, height-100, f"Giảng viên: {request.user.get_full_name() or request.user.username}")
        p.drawString(50, height-120, f"Ngày xuất: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Thống kê tổng quan
        bai_taps = BaiTap.objects.filter(nguoi_tao=request.user)
        tong_bai_nop = BaiLam.objects.filter(bai_tap__nguoi_tao=request.user).count()
        bai_da_cham = BaiLam.objects.filter(bai_tap__nguoi_tao=request.user, da_cham=True).count()
        
        y_pos = height - 160
        p.drawString(50, y_pos, f"Tổng số bài tập: {bai_taps.count()}")
        p.drawString(50, y_pos-20, f"Tổng bài nộp: {tong_bai_nop}")
        p.drawString(50, y_pos-40, f"Đã chấm: {bai_da_cham}")
        p.drawString(50, y_pos-60, f"Chưa chấm: {tong_bai_nop - bai_da_cham}")
        
        # Bảng chi tiết
        y_pos = height - 280
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y_pos, "STT")
        p.drawString(100, y_pos, "Tên bài tập")
        p.drawString(300, y_pos, "Loại")
        p.drawString(400, y_pos, "Tổng nộp")
        p.drawString(470, y_pos, "Đã chấm")
        
        p.setFont("Helvetica", 9)
        y_pos -= 20
        
        for i, bai_tap in enumerate(bai_taps[:20], 1):  # Giới hạn 20 bài
            bai_lams = BaiLam.objects.filter(bai_tap=bai_tap)
            da_cham = bai_lams.filter(da_cham=True)
            
            p.drawString(50, y_pos, str(i))
            p.drawString(100, y_pos, bai_tap.tieu_de[:25] + ('...' if len(bai_tap.tieu_de) > 25 else ''))
            p.drawString(300, y_pos, 'Lập trình' if bai_tap.loai_baitap == 'code' else 'Trắc nghiệm')
            p.drawString(400, y_pos, str(bai_lams.count()))
            p.drawString(470, y_pos, str(da_cham.count()))
            
            y_pos -= 15
            if y_pos < 100:  # Hết trang
                break
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        
        return response
        
    except Exception as e:
        messages.error(request, f'Lỗi xuất PDF: {str(e)}')
        return redirect('thong_ke_bao_cao')

